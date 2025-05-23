#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Merchant Report Generator Module

This module provides functionality for generating detailed Excel reports
for each merchant, including verification screenshots and merchant details.
"""

# Standard library imports
import os
import datetime
from typing import Dict, List, Any

# Third-party imports
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Local imports
from src.config.logging_config import get_logger

# Initialize logger
logger = get_logger(__name__)


def generate_merchant_report(
    merchant_data: Dict[str, Any],
    verification_results: Dict[str, Any],
    output_dir: str = "merchant_reports",
    include_screenshots: bool = True,
) -> str:
    """
    Generate a detailed Excel report for a specific merchant.

    Args:
        merchant_data: Dictionary containing merchant information
        verification_results: Dictionary with verification results including screenshots
        output_dir: Directory to save the report
        include_screenshots: Whether to include screenshots in the report

    Returns:
        Path to the generated report file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Generate filename
    merchant_id = str(merchant_data.get("merchant_id", "unknown"))
    merchant_name = str(merchant_data.get("merchant_name", "unknown")).replace(" ", "_")
    country = str(merchant_data.get("country", "unknown"))

    filename = f"{merchant_name}_{country}_{merchant_id}.xlsx"
    file_path = os.path.join(output_dir, filename)

    try:
        # Create a new workbook
        workbook = Workbook()

        # Create the verification results sheet (Sheet1)
        results_sheet = workbook.active
        results_sheet.title = "Verification Results"

        # Set up headers for verification results
        headers = [
            "Website URL",
            "Confidence Score",
            "Match Found",
            "Matching Text",
            "Screenshot",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = results_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Add verification data
        row_num = 2

        # Check if we have verification results
        if verification_results:
            url = verification_results.get("url", "N/A")
            confidence = verification_results.get("address_match_confidence", 0)
            verified = verification_results.get("verified", False)
            match_text = verification_results.get("address_match", "No match found")

            # Add data row
            results_sheet.cell(row=row_num, column=1).value = url
            results_sheet.cell(row=row_num, column=2).value = f"{confidence}%"
            results_sheet.cell(row=row_num, column=3).value = (
                "Yes" if verified else "No"
            )
            results_sheet.cell(row=row_num, column=4).value = match_text

            # Add screenshot if available and requested
            if include_screenshots:
                screenshot_path = verification_results.get("screenshot_path")
                if screenshot_path and os.path.exists(screenshot_path):
                    try:
                        # Add screenshot with adjusted size
                        img = Image(screenshot_path)

                        # Resize image to fit in the cell
                        max_width = 800
                        max_height = 600

                        # Calculate aspect ratio
                        width_ratio = (
                            max_width / img.width if img.width > max_width else 1
                        )
                        height_ratio = (
                            max_height / img.height if img.height > max_height else 1
                        )

                        # Use the smaller ratio to ensure image fits
                        ratio = min(width_ratio, height_ratio)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)

                        # Add to cell E2
                        results_sheet.add_image(img, "E2")

                        # Adjust row height to fit the image
                        results_sheet.row_dimensions[row_num].height = img.height
                    except Exception as e:
                        logger.warning(f"Failed to add screenshot: {str(e)}")
                        results_sheet.cell(
                            row=row_num, column=5
                        ).value = "(Screenshot unavailable)"
                else:
                    results_sheet.cell(
                        row=row_num, column=5
                    ).value = "(No screenshot available)"
        else:
            # No verification results
            results_sheet.cell(
                row=row_num, column=1
            ).value = "No verification data available"
            results_sheet.merge_cells(
                start_row=row_num, start_column=1, end_row=row_num, end_column=5
            )

        # Auto-adjust column widths for verification results
        for column in results_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Skip the screenshot column
            if column_letter == "E":
                results_sheet.column_dimensions[column_letter].width = 100
                continue

            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            results_sheet.column_dimensions[column_letter].width = adjusted_width

        # Create merchant details sheet (Sheet2)
        details_sheet = workbook.create_sheet(title="Merchant Details")

        # Add merchant details as key-value pairs
        details_sheet.cell(row=1, column=1).value = "Merchant Information"
        details_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        details_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        row_num = 2
        detail_fields = [
            ("Merchant ID", "merchant_id"),
            ("Merchant Name", "merchant_name"),
            ("Legal Name", "merchant_legal_name"),
            ("Industry", "industry"),
            ("Sub-Industry", "sub_industry"),
            ("Address", "address_line1"),
            ("Town", "town"),
            ("Postcode", "postcode"),
            ("Country", "country"),
            ("Verification Status", "verified"),
            ("Verification Confidence", "address_match_confidence"),
            ("Verification Date", "verification_date"),
        ]

        # Set today's date for verification date if not provided
        if verification_results and "verification_date" not in verification_results:
            verification_results["verification_date"] = (
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )

        # Add each merchant detail
        for label, field in detail_fields:
            # Get value from merchant_data or verification_results
            value = merchant_data.get(field, verification_results.get(field, "N/A"))

            # Format confidence as percentage
            if field == "address_match_confidence" and value != "N/A":
                value = f"{value}%"

            # Format verified as Yes/No
            if field == "verified" and value != "N/A":
                value = "Yes" if value else "No"

            # Add to sheet
            details_sheet.cell(row=row_num, column=1).value = label
            details_sheet.cell(row=row_num, column=1).font = Font(bold=True)
            details_sheet.cell(row=row_num, column=2).value = value

            row_num += 1

        # Auto-adjust column widths for merchant details
        for column in details_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            details_sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(file_path)
        logger.info(f"Generated merchant report: {file_path}")

        return file_path

    except Exception as e:
        logger.error(f"Error generating merchant report: {str(e)}")
        raise RuntimeError(f"Failed to generate merchant report: {str(e)}")


def generate_bulk_reports(
    merchants_df: pd.DataFrame,
    verification_results: Dict[str, Dict[str, Any]],
    output_dir: str = "merchant_reports",
) -> List[str]:
    """
    Generate individual reports for multiple merchants.

    Args:
        merchants_df: DataFrame containing merchant data
        verification_results: Dictionary mapping merchant_id to verification results
        output_dir: Directory to save the reports

    Returns:
        List of paths to generated report files
    """
    report_paths = []

    for _, merchant_row in merchants_df.iterrows():
        merchant_data = merchant_row.to_dict()
        merchant_id = str(merchant_data.get("merchant_id", ""))

        # Get verification results for this merchant
        merchant_results = verification_results.get(merchant_id, {})

        try:
            report_path = generate_merchant_report(
                merchant_data=merchant_data,
                verification_results=merchant_results,
                output_dir=output_dir,
            )
            report_paths.append(report_path)
        except Exception as e:
            logger.error(
                f"Failed to generate report for merchant {merchant_id}: {str(e)}"
            )

    return report_paths
